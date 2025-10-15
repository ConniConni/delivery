import configparser
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment
import datetime

# ロギング設定
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

# 工程番号と工程名のマッピング
PROCESS_MAP = {
    "030": "調査",
    "040": "設計",
    "050": "製造",
    "060": "UD作成",
    "070": "UD消化",
    "080": "SD作成",
    "090": "SD消化",
}

# ファイル種別
FILE_TYPE = {
    "1": "メイン資料",
    "2": "レビュー記録表",
    "3": "セルフチェック表",
}


def load_config(config_file="config.ini"):
    """設定ファイルを読み込む"""
    config = configparser.ConfigParser()
    config.read(config_file)
    return config


def create_empty_excel_file(type, title, project_name, item_name, file_path: Path):
    """空のExcelファイルを模したダミーファイルを作成する"""
    try:
        if not isinstance(type, str):
            raise TypeError("ファイル種別：typeは文字列で指定してください。")

        if type not in FILE_TYPE:
            raise ValueError(f"無効なファイル種別です: {type}")

    except (TypeError, ValueError) as e:
        logging.error(f"エラー: {e}")

    try:
        wb = Workbook()
        ws = wb.active
        # ws["A1"] = f"これはダミーのExcelファイルです。ファイル名: {file_path.name}"
        # ws.cell(row=1, column=1, value=f"これはダミーのExcelファイルです。ファイル名: {file_path.name}")

        if type == 1:
            ws["B5"] = f"{title}¥n{project_name}¥n{item_name}"
            # B5セルの書式を変更し、テキストの折り返しを有効にする。
            ws["B5"].alignment = Alignment(wrap_text=True)

        if type == 2:
            ws["B2"] = "セルフチェックリスト"
            ws["B3"] = title
            ws["B4"] = project_name
            ws["B5"] = item_name

        if type == 3:
            ws["B2"] = "レビュー記録表"
            ws["G2"] = f"{project_name}_{item_name}_{title} レビュー"

        wb.save(file_path)
        logging.debug(f"Created real Excel file: {file_path}")

    # Excelファイルの作成に失敗した場合はテキストファイルを作成
    except Exception as e:
        logging.warning(f"Failed to create dummy Excel file {file_path}: {e}")
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(f"Dummy Excel Content (Fallback) - {file_path.name}")


def create_sample_teams_structure(config):
    """サンプルのTeamsフォルダとファイル構造を生成する"""
    # config.iniから必要な情報のみを取得
    # KeyErrorが発生しないよう、get()メソッドを使用するか、事前にセクションとキーの存在を確認
    try:
        sample_teams_root = Path(config["Paths"]["sample_teams_root"])
        project_name = config["Project"]["project_name"]
        item_name = config["Project"]["item_name"]
    except KeyError as e:
        logging.error(f"config.iniに必須のセクションまたはキーがありません: {e}")
        logging.error(
            "config.iniには [Paths]セクションに 'sample_teams_root', [Project]セクションに 'project_name', 'item_name' が必要です。"
        )
        return

    # ルートフォルダの作成
    project_root = sample_teams_root / project_name / item_name
    project_root.mkdir(parents=True, exist_ok=True)
    logging.info(f"プロジェクトルートを作成しました: {project_root}")

    # 現在の日付情報に基づいた四半期フォルダ
    current_year = datetime.datetime.now().year
    current_quarter = (datetime.datetime.now().month - 1) // 3 + 1
    quarter_folder = f"{current_year}_{current_quarter}Q"

    teams_base_path = project_root / quarter_folder
    teams_base_path.mkdir(exist_ok=True)

    # 日付のサンプリング
    today = datetime.date.today()
    date1 = today - datetime.timedelta(days=7)  # 1週間前
    date2 = today - datetime.timedelta(days=5)
    date3 = today - datetime.timedelta(days=3)
    date4 = today - datetime.timedelta(days=1)

    for p_num, p_name in PROCESS_MAP.items():
        process_dir = teams_base_path / f"{p_num}.{p_name}"
        process_dir.mkdir(exist_ok=True)
        logging.info(f"  工程フォルダを作成: {process_dir.name}")

        # 主要成果物ファイル (工程フォルダ直下)
        main_excel_file_path = None
        type_main = 1

        if p_num == "030":  # 調査
            main_excel_file_path = (
                process_dir / f"調査検討書_{project_name}_{item_name}.xlsx"
            )
            title = config["title"]["research"]
            create_empty_excel_file(
                type_main, title, project_name, item_name, main_excel_file_path
            )
        elif p_num == "040":  # 設計
            main_excel_file_path = (
                process_dir / f"機能設計書_{project_name}_{item_name}.xlsx"
            )
            title = config["title"]["sys_design"]
            create_empty_excel_file(
                type_main, title, project_name, item_name, main_excel_file_path
            )
        elif p_num == "050":  # 製造 (Pythonファイル)
            (process_dir / f"xxx.py").touch()
        elif p_num == "060":  # UD作成
            main_excel_file_path = (
                process_dir / f"単体試験仕様書_{project_name}_{item_name}.xlsx"
            )
            title = config["title"]["unit_test_doc"]
            create_empty_excel_file(
                type_main, title, project_name, item_name, main_excel_file_path
            )
        elif p_num == "070":  # UD消化
            main_excel_file_path = (
                process_dir / f"単体試験成績書_{project_name}_{item_name}.xlsx"
            )
            title = config["title"]["unit_test_rst"]
            create_empty_excel_file(
                type_main, title, project_name, item_name, main_excel_file_path
            )
        elif p_num == "080":  # SD作成
            main_excel_file_path = (
                process_dir / f"結合試験仕様書_{project_name}_{item_name}.xlsx"
            )
            title = config["title"]["sys_test_doc"]
            create_empty_excel_file(
                type_main, title, project_name, item_name, main_excel_file_path
            )
        elif p_num == "090":  # SD消化
            rst_excel_file_path = (
                process_dir / f"結合試験成績書_{project_name}_{item_name}.xlsx"
            )
            rst_title = config["title"]["sys_test_rst"]
            create_empty_excel_file(
                type_main, rst_title, project_name, item_name, rst_excel_file_path
            )
            rep_excel_file_path = (
                process_dir / f"試験結果報告書_{project_name}_{item_name}.xlsx"
            )
            rep_title = config["title"]["test_rst_report"]
            create_empty_excel_file(
                type_main, rep_title, project_name, item_name, rep_excel_file_path
            )

        # 成果物/レビューフォルダ構造
        # 050.製造はレビューのみ
        type_review = 2

        # if p_num != "050" or p_num == "050":  # 全ての工程で成果物フォルダを作成
        #     results_dir = process_dir / "成果物"
        #     results_dir.mkdir(exist_ok=True)

        #     # 内部レビュー
        #     internal_review_dir = results_dir / "内部レビュー"
        #     internal_review_dir.mkdir(exist_ok=True)

        #     date_folder_int1 = internal_review_dir / date1.strftime("%Y%m%d")
        #     date_folder_int1.mkdir(exist_ok=True)
        #     if (
        #         main_excel_file_path and p_num != "090"
        #     ):  # 090はレビューフォルダにも複数のファイルがくる
        #         create_empty_excel_file(date_folder_int1 / main_excel_file_path.name)
        #     elif p_num == "090":
        #         create_empty_excel_file(
        #             date_folder_int1 / f"結合試験成績書_{project_name}_{item_name}.xlsx"
        #         )
        #         create_empty_excel_file(
        #             date_folder_int1 / f"試験結果報告書_{project_name}_{item_name}.xlsx"
        #         )

        #     create_empty_excel_file(
        #         date_folder_int1
        #         / f"レビューチェックリスト_{p_num}_社内_1回目_{project_name}_{item_name}.xlsx"
        #     )
        #     create_empty_excel_file(
        #         date_folder_int1
        #         / f"レビュー記録表_{p_name}_社内_1回目_{project_name}_{item_name}.xlsx"
        #     )

        # 030に2回目のレビューを追加
        # if p_num == "030":
        #     date_folder_int2 = internal_review_dir / date2.strftime("%Y%m%d")
        #     date_folder_int2.mkdir(exist_ok=True)
        #     create_empty_excel_file(
        #         date_folder_int2 / f"調査検討書_{project_name}_{item_name}.xlsx"
        #     )
        #     create_empty_excel_file(
        #         date_folder_int2
        #         / f"レビューチェックリスト_{p_num}_社内_2回目_{project_name}_{item_name}.xlsx"
        #     )
        #     create_empty_excel_file(
        #         date_folder_int2
        #         / f"レビュー記録表_{p_name}_社内_2回目_{project_name}_{item_name}.xlsx"
        #     )

        # 外部レビュー
        # external_review_dir = results_dir / "外部レビュー"
        # external_review_dir.mkdir(exist_ok=True)

        # 外部レビューがある工程のみ作成
        # if p_num in ["030", "040", "080", "090"]:
        #     date_folder_ext1 = external_review_dir / date3.strftime("%Y%m%d")
        #     date_folder_ext1.mkdir(exist_ok=True)
        #     if main_excel_file_path and p_num != "090":
        #         create_empty_excel_file(
        #             date_folder_ext1 / main_excel_file_path.name
        #         )
        #     elif p_num == "090":
        #         create_empty_excel_file(
        #             date_folder_ext1
        #             / f"結合試験成績書_{project_name}_{item_name}.xlsx"
        #         )
        #         create_empty_excel_file(
        #             date_folder_ext1
        #             / f"試験結果報告書_{project_name}_{item_name}.xlsx"
        #         )
        #     create_empty_excel_file(
        #         date_folder_ext1
        #         / f"レビューチェックリスト_{p_num}_社外_1回目_{project_name}_{item_name}.xlsx"
        #     )
        #     create_empty_excel_file(
        #         date_folder_ext1
        #         / f"レビュー記録表_{p_name}_社外_1回目_{project_name}_{item_name}.xlsx"
        #     )

        #     # 030に2回目の外部レビューを追加
        #     if p_num == "030":
        #         date_folder_ext2 = external_review_dir / date4.strftime("%Y%m%d")
        #         date_folder_ext2.mkdir(exist_ok=True)
        #         # 故意にファイル名を少し変えて、フェーズ1の収集ロジックのテストに使用
        #         create_empty_excel_file(
        #             date_folder_ext2 / f"調査B_{project_name}_{item_name}.xlsx"
        #         )
        #         create_empty_excel_file(
        #             date_folder_ext2
        #             / f"レビューチェックリスト_{p_num}_社外_2回目_{project_name}_{item_name}.xlsx"
        #         )
        #         create_empty_excel_file(
        #             date_folder_ext2
        #             / f"レビュー記録表_{p_name}_社外_2回目_{project_name}_{item_name}.xlsx"
        #         )

    logging.info("サンプルTeamsフォルダ構造の生成が完了しました。")


if __name__ == "__main__":
    config = load_config()
    # openpyxlがインストールされているかチェック
    try:
        import openpyxl

        logging.info(
            "openpyxlがインストールされています。実際に近いExcelファイルを生成します。"
        )
    except ImportError:
        logging.warning(
            "openpyxlがインストールされていません。ダミーのテキストファイルが.xlsx拡張子で生成されます。"
        )
        logging.warning(
            "実際のExcelファイルを作成するには 'pip install openpyxl' を実行してください。"
        )

    create_sample_teams_structure(config)
