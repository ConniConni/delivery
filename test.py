import configparser
import logging
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment
import datetime

# ロギング設定
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)

# ファイルタイプの定数を定義
FILE_TYPE_MAIN = 1
FILE_TYPE_REVIEW_CHECKLIST = 2
FILE_TYPE_REVIEW_MINUTES = 3
FILE_TYPE_PYTHON = 4  # Pythonファイル用

# 各工程での主要成果物ファイル情報
PROCESS_FILES = {
    "030": {"main": "調査検討書", "title_key": "research"},
    "040": {"main": "機能設計書", "title_key": "sys_design"},
    "050": {"main": "xxx", "type": FILE_TYPE_PYTHON},  # Pythonファイルは特殊
    "060": {"main": "単体試験仕様書", "title_key": "unit_test_doc"},
    "070": {"main": "単体試験成績書", "title_key": "unit_test_rst"},
    "080": {"main": "結合試験仕様書", "title_key": "sys_test_doc"},
    "090": [
        {"main": "結合試験成績書", "title_key": "sys_test_rst"},
        {"main": "試験結果報告書", "title_key": "test_rst_report"},
    ],
}

# 工程番号と工程名のマッピング
PROCESS_MAP = {
    "030": "調査",
    "040": "設計",
    "050": "製造",
    "060": "UD作成",
    "070": "UD消化",
    "080": "SD作成",
    "090": "SD消化",
}

# ファイル種別 (create_dummy_excel_file用)
FILE_TYPE = {
    1: "メイン資料",
    2: "レビューチェックリスト",  # 以前は「レビュー記録表」でしたが、定数名に合わせて修正
    3: "レビュー記録表",  # 以前は「セルフチェック表」でしたが、定数名に合わせて修正
}


def load_config(config_file="config.ini"):
    """設定ファイルを読み込む"""
    config = configparser.ConfigParser()
    config.read(config_file)
    return config


def create_dummy_excel_file(type, title, project_name, item_name, file_path: Path):
    """空のExcelファイルを模したダミーファイルを作成する"""
    try:
        if not isinstance(type, int):
            raise TypeError("ファイル種別：typeは整数で指定してください。")

        if type not in FILE_TYPE:
            raise ValueError(f"無効なファイル種別です: {type}")

    except (TypeError, ValueError) as e:
        logging.error(f"エラー: {e}")
        return False  # エラー時はここで処理を中断し、Falseを返す

    try:
        wb = Workbook()
        ws = wb.active

        if type == FILE_TYPE_MAIN:  # メイン資料
            ws["B5"] = f"{title}\n{project_name}\n{item_name}"  # '\n' で改行
            ws["B5"].alignment = Alignment(wrap_text=True)
            ws["A1"] = f"これはメイン資料です。ファイル名: {file_path.name}"

        elif type == FILE_TYPE_REVIEW_CHECKLIST:  # レビューチェックリスト
            ws["B2"] = "レビューチェックリスト"
            ws["B3"] = title  # ここには "レビューチェックリスト" という文字列が来る
            ws["B4"] = project_name
            ws["B5"] = item_name
            ws["A1"] = f"これはレビューチェックリストです。ファイル名: {file_path.name}"

        elif type == FILE_TYPE_REVIEW_MINUTES:  # レビュー記録表
            ws["B2"] = "レビュー記録表"
            ws["G2"] = (
                f"{project_name}_{item_name}_{title} レビュー"  # ここには "レビュー記録表" という文字列が来る
            )
            ws["A1"] = f"これはレビュー記録表です。ファイル名: {file_path.name}"

        else:  # 未定義のタイプの場合のフォールバック
            ws["A1"] = (
                f"不明なファイルタイプ ({type}) のダミーExcelです。ファイル名: {file_path.name}"
            )

        wb.save(file_path)
        logging.info(
            f"Created real Excel file: {file_path.name}"
        )  # INFOレベルで作成を確認
        return True  # 成功したことを示す

    # Excelファイルの作成に失敗した場合はテキストファイルを作成
    except Exception as e:
        logging.warning(f"Failed to create dummy Excel file {file_path}: {e}")
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(f"Dummy Excel Content (Fallback) - {file_path.name}\n")
            f.write(f"Error: {e}\n")
        return False  # 失敗したことを示す


def create_file_with_title(
    file_type: int,
    base_dir: Path,
    file_prefix: str,
    config: configparser.ConfigParser,
    project_name: str,
    item_name: str,
    specific_title_key: str = None,  # config[title] から取得するキー
    file_extension: str = ".xlsx",
) -> Path | None:
    """
    指定された情報に基づいてファイルを作成するヘルパー関数。
    Pythonファイルの場合は .py 拡張子で空ファイルを作成し、それ以外はExcelダミーファイルを作成。
    """
    # base_dirが存在しない場合は作成
    base_dir.mkdir(parents=True, exist_ok=True)

    if file_type == FILE_TYPE_PYTHON:
        file_path = base_dir / f"{file_prefix}.py"
        file_path.touch()
        logging.debug(f"Created Python file: {file_path.name}")
        return file_path
    else:
        # titleはconfigから取得するか、直接prefixをタイトルとして使用
        title_key = (
            specific_title_key
            if specific_title_key
            else file_prefix.replace("_", "").lower()
        )
        logging.debug(f"Resolved title_key for config lookup: '{title_key}'")

        # config['title'] に 'review_checklist' や 'review_minutes' がある前提
        # もし特定のファイル名のタイトル（例: "調査検討書"）を使いたい場合は
        # specific_title_key をそれに合わせて設定する必要がある。
        title_from_config = config["title"].get(
            title_key, file_prefix
        )  # configにない場合のフォールバック
        logging.debug(f"Title from config: '{title_from_config}'")

        file_name = f"{file_prefix}_{project_name}_{item_name}{file_extension}"
        file_path = base_dir / file_name

        success = create_dummy_excel_file(
            file_type, title_from_config, project_name, item_name, file_path
        )
        if success:
            return file_path
        else:
            return None


# ヘルパー関数: レビュー関連ファイルを作成する
def create_review_files(
    base_folder: Path,
    date_obj: datetime.date,
    suffix: str,
    config: configparser.ConfigParser,
    project_name: str,
    item_name: str,
    p_num: str,
    p_name: str,
    main_file_path_stems: list[
        str
    ],  # メイン成果物のファイル名プレフィックスのリスト (例: "調査検討書", "機能設計書")
):
    """レビュー関連ファイルを作成するヘルパー関数"""
    date_folder = base_folder / date_obj.strftime("%Y%m%d")
    date_folder.mkdir(exist_ok=True)
    logging.info(f"    日付フォルダを作成: {date_folder.name}")

    # 主要成果物ファイルをコピーまたは再作成
    for main_stem in main_file_path_stems:
        # title_keyの推測ロジックを強化または明示的に指定
        # ここでは、main_stemが"調査検討書"のような場合を想定し、その一部をタイトルキーとして使う
        # 必要に応じて、PROCESS_FILESから該当するタイトルキーを取得するロジックを追加
        inferred_title_key = (
            main_stem.replace("書", "").lower()
            if "書" in main_stem
            else main_stem.lower()
        )
        if main_stem.startswith("結合試験成績書"):  # 結合試験成績書は"sys_test_rst"
            inferred_title_key = "sys_test_rst"
        elif main_stem.startswith(
            "試験結果報告書"
        ):  # 試験結果報告書は"test_rst_report"
            inferred_title_key = "test_rst_report"

        create_file_with_title(
            FILE_TYPE_MAIN,
            date_folder,
            main_stem,  # 拡張子を除いたファイル名プレフィックス
            config,
            project_name,
            item_name,
            specific_title_key=inferred_title_key,  # 推測したタイトルキーを渡す
        )

    # レビューチェックリストとレビュー記録表を作成
    create_file_with_title(
        FILE_TYPE_REVIEW_CHECKLIST,
        date_folder,
        f"レビューチェックリスト_{p_num}_{suffix}",
        config,
        project_name,
        item_name,
        specific_title_key="review_checklist",
    )
    create_file_with_title(
        FILE_TYPE_REVIEW_MINUTES,
        date_folder,
        f"レビュー記録表_{p_name}_{suffix}",
        config,
        project_name,
        item_name,
        specific_title_key="review_minutes",
    )
    return date_folder  # 作成した日付フォルダを返す


def create_sample_teams_structure(config):
    """サンプルのTeamsフォルダとファイル構造を生成する"""
    try:
        sample_teams_root = Path(config["Paths"]["sample_teams_root"])
        project_name = config["Project"]["project_name"]
        item_name = config["Project"]["item_name"]
    except KeyError as e:
        logging.error(f"config.iniに必須のセクションまたはキーがありません: {e}")
        logging.error(
            "config.iniには [Paths]セクションに 'sample_teams_root', [Project]セクションに 'project_name', 'item_name' が必要です。"
        )
        return

    # ルートフォルダの作成
    project_root = sample_teams_root / project_name / item_name
    project_root.mkdir(parents=True, exist_ok=True)
    logging.info(f"プロジェクトルートを作成しました: {project_root}")

    # 現在の日付情報に基づいた四半期フォルダ
    current_year = datetime.datetime.now().year
    current_quarter = (datetime.datetime.now().month - 1) // 3 + 1
    quarter_folder = f"{current_year}_{current_quarter}Q"

    teams_base_path = project_root / quarter_folder
    teams_base_path.mkdir(exist_ok=True)
    logging.info(f"四半期フォルダを作成しました: {teams_base_path}")

    # 日付のサンプリング
    today = datetime.date.today()
    DATE_SAMPLES = {
        "date1": today - datetime.timedelta(days=7),  # 1週間前 (内部レビュー1回目)
        "date2": today - datetime.timedelta(days=5),  # 内部レビュー2回目
        "date3": today - datetime.timedelta(days=3),  # 外部レビュー1回目
        "date4": today - datetime.timedelta(days=1),  # 外部レビュー2回目
    }

    # 各工程のループ
    for p_num, p_name in PROCESS_MAP.items():
        process_dir = teams_base_path / f"{p_num}.{p_name}"
        process_dir.mkdir(exist_ok=True)
        logging.info(f"  工程フォルダを作成: {process_dir.name}")

        # その工程で作成される主要成果物ファイル名を収集するリスト
        # このリストはレビュー時にも利用される
        main_file_stem_for_review: list[str] = []

        # 主要成果物ファイルの作成
        process_file_info = PROCESS_FILES.get(p_num)

        if isinstance(process_file_info, dict):  # 単一ファイルの工程
            # ここでPythonファイル（050）かどうかをチェックする
            file_type_to_use = process_file_info.get("type", FILE_TYPE_MAIN)

            if file_type_to_use == FILE_TYPE_PYTHON:
                file_path = create_file_with_title(
                    file_type_to_use,
                    process_dir,
                    process_file_info["main"],
                    config,
                    project_name,
                    item_name,
                    specific_title_key=None,  # Pythonファイルなのでtitle_keyは不要
                )
            else:
                file_path = create_file_with_title(
                    file_type_to_use,
                    process_dir,
                    process_file_info["main"],
                    config,
                    project_name,
                    item_name,
                    process_file_info[
                        "title_key"
                    ],  # 通常のExcelファイルの場合はtitle_keyが必要
                )
            if file_path:
                main_file_stem_for_review.append(file_path.stem)

        elif isinstance(process_file_info, list):  # 複数ファイルの工程 (例: 090)
            for file_data in process_file_info:
                file_path = create_file_with_title(
                    file_data.get("type", FILE_TYPE_MAIN),
                    process_dir,
                    file_data["main"],
                    config,
                    project_name,
                    item_name,
                    file_data["title_key"],
                )
                if file_path:
                    main_file_stem_for_review.append(file_path.stem)
        # else: プロセスにファイルが定義されていない場合は何もしない (例: 050でPythonファイル以外ない場合など)

        # 成果物/レビューフォルダ構造の作成
        results_dir = process_dir / "成果物"
        results_dir.mkdir(exist_ok=True)
        logging.info(f"    成果物フォルダを作成: {results_dir.name}")

        # 内部レビューフォルダの作成
        internal_review_dir = results_dir / "内部レビュー"
        internal_review_dir.mkdir(exist_ok=True)
        logging.info(f"    内部レビューフォルダを作成: {internal_review_dir.name}")

        # 内部レビュー1回目 (全ての工程で作成)
        create_review_files(
            internal_review_dir,
            DATE_SAMPLES["date1"],
            "社内_1回目",
            config,
            project_name,
            item_name,
            p_num,
            p_name,
            main_file_stem_for_review,  # その工程の主要成果物ファイルを渡す
        )

        # 030に2回目の内部レビューを追加
        if p_num == "030":
            create_review_files(
                internal_review_dir,
                DATE_SAMPLES["date2"],
                "社内_2回目",
                config,
                project_name,
                item_name,
                p_num,
                p_name,
                main_file_stem_for_review,  # 030の主要成果物ファイルを渡す
            )

        # 外部レビューフォルダの作成
        external_review_dir = results_dir / "外部レビュー"
        external_review_dir.mkdir(exist_ok=True)
        logging.info(f"    外部レビューフォルダを作成: {external_review_dir.name}")

        # 外部レビューがある工程のみ作成
        if p_num in ["030", "040", "080", "090"]:
            # 外部レビュー1回目
            created_ext1_folder = create_review_files(
                external_review_dir,
                DATE_SAMPLES["date3"],
                "社外_1回目",
                config,
                project_name,
                item_name,
                p_num,
                p_name,
                main_file_stem_for_review,  # その工程の主要成果物ファイルを渡す
            )

            # 030に2回目の外部レビューを追加
            if p_num == "030":
                # 故意にファイル名を少し変える部分は、個別の create_file_with_title で対応
                # ここでは "調査B" という新しいプレフィックスでファイルを作成
                create_file_with_title(
                    FILE_TYPE_MAIN,
                    created_ext1_folder,  # 既存の外部レビュー1回目のフォルダを使用
                    "調査B",  # 変更後のファイル名プレフィックス
                    config,
                    project_name,
                    item_name,
                    specific_title_key="research",  # 元のタイトルキーを使用
                )
                # 2回目の外部レビューのチェックリストと記録表は通常通り作成
                create_review_files(
                    external_review_dir,
                    DATE_SAMPLES["date4"],
                    "社外_2回目",
                    config,
                    project_name,
                    item_name,
                    p_num,
                    p_name,
                    [],  # ここでは追加のメイン成果物ファイルは作成しない (調査Bは上記で作成済み)
                )

    logging.info("サンプルTeamsフォルダ構造の生成が完了しました。")


if __name__ == "__main__":
    # config.ini の内容が適切に設定されていることを確認してください。
    # 特に [title] セクションに "research", "sys_design", "review_checklist" など
    # 各ファイルタイプに対応するキーが存在すること。
    # [Paths] の sample_teams_root も忘れずに設定してください。
    # 例: sample_teams_root = C:\temp\GeneratedTeamsProjects

    config = load_config("config.ini")

    # 必須セクションの存在確認
    required_sections = ["Paths", "Project", "title"]
    for section in required_sections:
        if not config.has_section(section):
            logging.error(
                f"config.iniに必須のセクション '[{section}]' がありません。デバッグを中断します。"
            )
            exit(1)  # エラー終了

    # create_sample_teams_structure 関数を呼び出す
    create_sample_teams_structure(config)

    logging.info("\n--- 全工程の生成が完了しました。---")
    try:
        root_path = (
            Path(config["Paths"]["sample_teams_root"])
            / config["Project"]["project_name"]
            / config["Project"]["item_name"]
        )
        logging.info(f"生成されたフォルダ構造は '{root_path}' 以下にあります。")
    except KeyError:
        pass  # configの読み込みでエラーが出ている場合はパスが存在しないため何もしない
